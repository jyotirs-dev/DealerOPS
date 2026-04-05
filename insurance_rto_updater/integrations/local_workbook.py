"""
local_workbook.py – Local Excel workbook read / write adapter
==============================================================
Loads a local workbook, finds the first worksheet with the fixed sales-sheet
headers, exposes its data to the processing pipeline, and applies the
resulting write plan back into the workbook.
"""
from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from insurance_rto_updater.domain.normalization import normalize_text
from insurance_rto_updater.models import SheetData, SheetWritePlan


FIXED_CUSTOMER_HEADER = "Contact Name"
FIXED_INSURANCE_HEADER = "Insurance"
FIXED_RTO_HEADER = "(RTO+ Agent fee 500)"
_REQUIRED_HEADERS = (
    FIXED_CUSTOMER_HEADER,
    FIXED_INSURANCE_HEADER,
    FIXED_RTO_HEADER,
)


def _trim_trailing_empty(values: list[str]) -> list[str]:
    """Drop trailing empty cells from a header-like row."""
    trimmed = list(values)
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return trimmed


def _cell_to_text(value: Any) -> str:
    """Convert a workbook cell value into the trimmed text used by the pipeline."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value).strip()


def _cell_to_json(value: Any) -> Any:
    """Convert a workbook cell value into a JSON-serializable primitive."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


class LocalWorkbookAdapter:
    """
    Read and write the uploaded Excel workbook used by the React workflow.

    The adapter targets the first worksheet whose first row contains the fixed
    customer/insurance/RTO headers after normalization.
    """

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self._workbook: Workbook | None = None
        self._worksheet: Worksheet | None = None

    def _load(self) -> Workbook:
        if self._workbook is None:
            if not self.workbook_path.exists():
                raise ValueError(f"Workbook not found: {self.workbook_path}")
            try:
                self._workbook = load_workbook(
                    self.workbook_path,
                    keep_vba=self.workbook_path.suffix.lower() == ".xlsm",
                )
            except Exception as exc:
                raise ValueError(f"Failed to read workbook: {exc}") from exc
        return self._workbook

    def _header_row(self, worksheet: Worksheet) -> list[str]:
        values = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        return _trim_trailing_empty([_cell_to_text(value) for value in values])

    def _resolve_worksheet(self) -> Worksheet:
        if self._worksheet is not None:
            return self._worksheet

        workbook = self._load()
        required = {normalize_text(value) for value in _REQUIRED_HEADERS}

        for worksheet in workbook.worksheets:
            header_row = self._header_row(worksheet)
            available = {
                normalize_text(cell)
                for cell in header_row
                if normalize_text(cell)
            }
            if required.issubset(available):
                self._worksheet = worksheet
                return worksheet

        raise ValueError(
            "Workbook must contain a worksheet with headers: "
            f"{', '.join(_REQUIRED_HEADERS)}"
        )

    def _serialize_rows(self, worksheet: Worksheet) -> list[list[Any]]:
        rows: list[list[Any]] = []
        max_col = max(1, len(self._header_row(worksheet)))
        for values in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            max_col=max_col,
            values_only=True,
        ):
            rows.append([_cell_to_json(value) for value in values])
        return rows

    def load_sheet_data(self) -> SheetData:
        worksheet = self._resolve_worksheet()
        header_row = self._header_row(worksheet)
        data_rows = [
            [_cell_to_text(value) for value in values]
            for values in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                max_col=max(1, len(header_row)),
                values_only=True,
            )
        ]

        return SheetData(
            spreadsheet_id=self.workbook_path.stem,
            spreadsheet_url=self.workbook_path.name,
            sheet_title=worksheet.title,
            header_row=header_row,
            data_rows=data_rows,
        )

    def apply_write_plan(
        self,
        sheet_title: str,
        write_plan: SheetWritePlan,
    ) -> None:
        workbook = self._load()
        try:
            worksheet = workbook[sheet_title]
        except KeyError as exc:
            raise ValueError(f"Worksheet '{sheet_title}' not found.") from exc

        if (
            write_plan.clear_to_row >= write_plan.clear_from_row
            and write_plan.clear_columns
        ):
            for row_idx in range(
                write_plan.clear_from_row, write_plan.clear_to_row + 1
            ):
                for col_idx in write_plan.clear_columns:
                    worksheet.cell(row=row_idx, column=col_idx).value = None

        for item in write_plan.value_updates:
            worksheet.cell(
                row=item.row_index,
                column=item.col_index,
            ).value = item.value

    def sheet_preview(
        self,
        sheet_title: str | None = None,
    ) -> tuple[list[str], list[list[Any]]]:
        worksheet = self._resolve_worksheet()
        if sheet_title is not None and worksheet.title != sheet_title:
            workbook = self._load()
            try:
                worksheet = workbook[sheet_title]
            except KeyError as exc:
                raise ValueError(f"Worksheet '{sheet_title}' not found.") from exc

        return self._header_row(worksheet), self._serialize_rows(worksheet)

    def save_copy(self, destination: Path) -> Path:
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = self._load()
        workbook.save(destination)
        return destination
