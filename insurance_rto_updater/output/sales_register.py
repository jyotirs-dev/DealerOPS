"""
sales_register.py – Vehicle Sales Register generator
=====================================================
Transforms a raw OEM / DMS invoice export (.xlsx) into a formatted
Vehicle Sales Register workbook with:

  • Auto-filled columns from raw data  (A, B, C, D, E, F, G, H, J, O, Z)
  • Formula columns                     (K, P, Q, S, T, U, W, X)
  • Blank columns for manual entry      (I, L, M, N, R, V, Y, AA, AB)

The output is a standalone .xlsx file ready for the dealership to open,
fill in the manual columns, and let the formulas compute the rest.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from insurance_rto_updater.models import SalesRegisterResult


# ── Constants ─────────────────────────────────────────────────────────────────

DEALER_NAME = "SHREE DURGA DARSHAN AUTOMOBILES"

HEADERS = [
    "Invoice Date",
    "Invoice No.",
    "Contact Name",
    "Contact \nMobile No.",
    "Contact \nAddress",
    "Model Name",
    "Color",
    "VIN",
    "Discount \nPre",
    "Total\n Invoice",
    "Ex \nShowroom",
    "(RTO+\n Agent fee 500)",
    "Insurance",
    "Accesories",
    "Discount",
    "Expected \nOn Road Price",
    "Actual On\nRoad Price",
    "Down \nPayment",
    "Filed DP\n(to Bank)",
    "Our \nComission",
    "Agent \nComission",
    "FINANCE \nDISBURSEMENT",
    "Expected \nGAP",
    "Actual \nGap",
    "Outstanding",
    "Financer Name (HP With)",
    "Remarks",
    "Status",
]

MANUAL_COLUMNS = [
    "I  – Discount Pre",
    "L  – RTO + Agent Fee",
    "M  – Insurance premium",
    "N  – Accessories",
    "R  – Down Payment",
    "V  – Finance Disbursement",
    "Y  – Outstanding",
    "AA – Remarks",
    "AB – Status",
]

FONT_SIZE = 9
GST_RATE = 0.18
EX_SHOWROOM_DEDUCTION = 1099.72
CHAR_WIDTH = 1.1


# ── Styling helpers ───────────────────────────────────────────────────────────

def _thin_border() -> Border:
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style(cell) -> None:  # type: ignore[no-untyped-def]
    cell.font = Font(bold=True, size=FONT_SIZE)
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True,
    )


def _data_style(cell) -> None:  # type: ignore[no-untyped-def]
    cell.font = Font(size=FONT_SIZE)
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ── Raw-column extractor ─────────────────────────────────────────────────────

def _safe_value(row: pd.Series, col_name: str):  # type: ignore[type-arg]
    """Return the column value or None when missing / NaN."""
    val = row.get(col_name)
    if val is None or pd.isna(val):
        return None
    return val


# ── Public API ────────────────────────────────────────────────────────────────

def generate_sales_register(
    raw_path: Path,
    output_path: Path,
) -> SalesRegisterResult:
    """
    Read a raw invoice export and produce a styled Vehicle Sales Register.

    Parameters
    ----------
    raw_path:
        Path to the raw .xlsx file exported from the OEM / DMS system.
    output_path:
        Destination for the generated workbook.

    Returns
    -------
    SalesRegisterResult with the output path, row count, and month/year.
    """
    df = pd.read_excel(raw_path, sheet_name=0)
    df.columns = [c.strip() for c in df.columns]

    # Keep only invoiced rows.
    df = df[df["Invoice Status"] == "Invoiced"].reset_index(drop=True)

    # Detect month-year from the first invoice date.
    month_year = ""
    if not df.empty:
        first_date = pd.to_datetime(
            df["Invoice Date"].dropna().iloc[0], errors="coerce",
        )
        if pd.notna(first_date):
            month_year = first_date.strftime("%b%Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicle Sales Register"

    # ── Header row ────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30.0
    for c_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        _header_style(cell)
        cell.border = _thin_border()

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, row_data in df.iterrows():
        r = int(i) + 2  # type: ignore[arg-type]  # Excel row (1=header)

        # A: Invoice Date
        ws[f"A{r}"] = _safe_value(row_data, "Invoice Date")
        ws[f"A{r}"].number_format = "d mmmm yyyy"

        # B: Invoice Number
        ws[f"B{r}"] = _safe_value(row_data, "Invoice Number")

        # C: Contact Name
        ws[f"C{r}"] = _safe_value(row_data, "Contact Name")

        # D: Contact Mobile Number
        mobile = _safe_value(row_data, "Contact Mobile Number")
        ws[f"D{r}"] = int(mobile) if mobile else None

        # E: Contact Address
        ws[f"E{r}"] = _safe_value(row_data, "Contact Address")

        # F: Model Name
        ws[f"F{r}"] = _safe_value(row_data, "Model Name")

        # G: Color
        ws[f"G{r}"] = _safe_value(row_data, "Color")

        # H: VIN
        ws[f"H{r}"] = _safe_value(row_data, "VIN")

        # I: Discount Pre — manual entry (blank)

        # J: Total Invoice Amount
        ws[f"J{r}"] = _safe_value(row_data, "Total Invoice Amount")
        ws[f"J{r}"].number_format = "0.00"

        # K: Ex Showroom = J - 1099.72
        ws[f"K{r}"] = f"=J{r}-{EX_SHOWROOM_DEDUCTION}"
        ws[f"K{r}"].number_format = "0.00"

        # L: RTO + Agent fee — manual entry (blank)
        # M: Insurance — manual entry (blank)
        # N: Accessories — manual entry (blank)

        # O: Discount = Pre Vat Discount + 18% GST (0 if no discount)
        pre_vat = _safe_value(row_data, "Pre Vat Discount")
        if pre_vat and float(pre_vat) > 0:
            ws[f"O{r}"] = round(float(pre_vat) * (1 + GST_RATE), 2)
        else:
            ws[f"O{r}"] = None

        # P: Expected On Road Price = K+L+M+N-O+200
        ws[f"P{r}"] = f"=K{r}+L{r}+M{r}+N{r}-O{r}+200"
        ws[f"P{r}"].number_format = "0.00"

        # Q: Actual On Road Price = J+L+M-O
        ws[f"Q{r}"] = f"=J{r}+L{r}+M{r}-O{r}"
        ws[f"Q{r}"].number_format = "0.00"

        # R: Down Payment — manual entry (blank)

        # S: Filed DP to Bank = IF(ISTEXT(Z), R-3500, R)
        ws[f"S{r}"] = f"=IF(ISTEXT(Z{r}),(R{r}-3500),R{r})"
        ws[f"S{r}"].number_format = "0.00"

        # T: Our Commission = R-S-U
        ws[f"T{r}"] = f"=R{r}-S{r}-U{r}"
        ws[f"T{r}"].number_format = "0.00"

        # U: Agent Commission = IF(ISTEXT(Z), 1000, 0)
        ws[f"U{r}"] = f"=IF(ISTEXT(Z{r}),1000,0)"
        ws[f"U{r}"].number_format = "0.00"

        # V: Finance Disbursement — manual entry (blank)

        # W: Expected GAP = P-S-V
        ws[f"W{r}"] = f"=P{r}-S{r}-V{r}"
        ws[f"W{r}"].number_format = "0.00"

        # X: Actual Gap = Q-S-V
        ws[f"X{r}"] = f"=Q{r}-S{r}-V{r}"
        ws[f"X{r}"].number_format = "0.00"

        # Y: Outstanding — manual entry (blank)

        # Z: Financer Name
        ws[f"Z{r}"] = _safe_value(row_data, "Financer Name")

        # AA: Remarks — manual entry (blank)
        # AB: Status — manual entry (blank)

        # Apply consistent styling to every cell in the row.
        for c_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c_idx)
            _data_style(cell)
            cell.border = _thin_border()

    # ── Auto-fit column widths ────────────────────────────────────────────
    for c_idx in range(1, len(HEADERS) + 1):
        col_letter = get_column_letter(c_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=c_idx, max_col=c_idx):
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                lines = str(cell.value).split("\n")
                cell_len = max(len(line) for line in lines)
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = max(
            max_len * CHAR_WIDTH + 2, 8,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return SalesRegisterResult(
        output_path=output_path,
        rows_written=len(df),
        month_year=month_year,
        manual_columns=list(MANUAL_COLUMNS),
    )
