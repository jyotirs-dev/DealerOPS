from __future__ import annotations

import io
import pandas as pd
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import app as app_module
from insurance_rto_updater.integrations.local_workbook import (
    FIXED_CUSTOMER_HEADER,
    FIXED_INSURANCE_HEADER,
    FIXED_RTO_HEADER,
)

INSURANCE_REPORT_TEXT = """MIS BUSINESS REPORT USER WISE

S.No.
User Name
Policy
Type
Policy Number
Customer
Name
Start
Date
OD
Premium
NCB
ND
Cover
RTI
Cover
RSA
addons
Gross
Premium
1.
MAHENDRA61835
N
993792623750035786
VIRENDRA
VALIYA
2/2/2026
8:00:28
PM
851
0
YES
NO
NO
5548
2.
MAHENDRA61835
N
993792623750039494 IMAM SHAH
2/6/2026
5:12:13
PM
802
0
YES
NO
NO
5491
Total
1653
0
11039
"""


class ApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.upload_root = Path(self.temp_dir.name) / "uploads"
        self.output_root = Path(self.temp_dir.name) / "outputs"
        self.upload_patch = patch.object(app_module, "UPLOAD_ROOT", self.upload_root)
        self.output_patch = patch.object(app_module, "OUTPUT_ROOT", self.output_root)
        self.upload_patch.start()
        self.output_patch.start()
        self.addCleanup(self.upload_patch.stop)
        self.addCleanup(self.output_patch.stop)
        app_module._ensure_dirs()
        app_module.app.config["TESTING"] = True
        self.client = app_module.app.test_client()

    def _workbook_bytes(
        self,
        existing_rto: int | str = "",
        customer_rows: list[list[str | int]] | None = None,
    ) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Vehicle Sales Register"
        worksheet.append(
            [
                FIXED_CUSTOMER_HEADER,
                FIXED_INSURANCE_HEADER,
                FIXED_RTO_HEADER,
            ]
        )
        rows = customer_rows or [
            ["Ramesh Kumar", "", ""],
            ["Suresh Sharma", "", existing_rto],
        ]
        for row in rows:
            worksheet.append(row)
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @patch("insurance_rto_updater.orchestration.pipeline.extract_text_from_file")
    def test_process_endpoint_updates_and_returns_workbook(
        self,
        mocked_extract,
    ) -> None:
        def fake_extract(path: Path) -> str:
            if "insurance" in path.name:
                return "Insured: Ramesh Kumar\nGrand Total: 5400"
            return "Received From: Suresh Sharma\nGrand Total: 3200"

        mocked_extract.side_effect = fake_extract

        response = self.client.post(
            "/api/process",
            data={
                "workbook": (
                    io.BytesIO(self._workbook_bytes()),
                    "sales.xlsx",
                ),
                "insurance_files": (
                    io.BytesIO(b"insurance"),
                    "insurance-one.pdf",
                ),
                "rto_files": (
                    io.BytesIO(b"rto"),
                    "rto-one.pdf",
                ),
                "clear_existing": "1",
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        assert payload is not None

        self.assertEqual(payload["sheetTitle"], "Vehicle Sales Register")
        self.assertEqual(payload["summary"]["billsProcessed"], 2)
        self.assertEqual(payload["summary"]["billsUpdated"], 2)
        self.assertEqual(payload["reviewRows"], [])
        self.assertTrue(payload["downloadUrl"].endswith("sales_updated.xlsx"))
        self.assertTrue(payload["reviewCsvUrl"].endswith("review_conflicts.csv"))
        self.assertEqual(payload["rows"][0][1], 5400.0)
        self.assertEqual(payload["rows"][1][2], 3700.0)

        review_response = self.client.get(payload["reviewCsvUrl"])
        self.assertEqual(review_response.status_code, 200)

        download_response = self.client.get(payload["downloadUrl"])
        self.assertEqual(download_response.status_code, 200)

        workbook = load_workbook(io.BytesIO(download_response.data))
        worksheet = workbook[payload["sheetTitle"]]
        self.assertEqual(worksheet.cell(row=2, column=2).value, 5400.0)
        self.assertEqual(worksheet.cell(row=3, column=3).value, 3700.0)

    @patch("insurance_rto_updater.orchestration.pipeline.extract_text_from_file")
    def test_process_endpoint_updates_multiple_rows_from_insurance_report(
        self,
        mocked_extract,
    ) -> None:
        def fake_extract(path: Path) -> str:
            if "insurance" in path.name:
                return INSURANCE_REPORT_TEXT
            return "Received From: Suresh Sharma\nGrand Total: 3200"

        mocked_extract.side_effect = fake_extract

        response = self.client.post(
            "/api/process",
            data={
                "workbook": (
                    io.BytesIO(
                        self._workbook_bytes(
                            customer_rows=[
                                ["VIRENDRA VALIYA", "", ""],
                                ["IMAM SHAH", "", ""],
                                ["Suresh Sharma", "", ""],
                            ]
                        )
                    ),
                    "sales.xlsx",
                ),
                "insurance_files": (
                    io.BytesIO(b"insurance-report"),
                    "insurance-report.pdf",
                ),
                "rto_files": (
                    io.BytesIO(b"rto"),
                    "rto-one.pdf",
                ),
                "clear_existing": "1",
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        assert payload is not None

        self.assertEqual(payload["summary"]["billsProcessed"], 3)
        self.assertEqual(payload["summary"]["billsUpdated"], 3)
        self.assertEqual(payload["summary"]["rowsUpdated"], 3)
        self.assertEqual(payload["reviewRows"], [])
        self.assertEqual(payload["rows"][0][1], 5548.0)
        self.assertEqual(payload["rows"][1][1], 5491.0)
        self.assertEqual(payload["rows"][2][2], 3700.0)

    @patch("insurance_rto_updater.orchestration.pipeline.extract_text_from_file")
    def test_process_endpoint_returns_review_rows_with_reason(
        self,
        mocked_extract,
    ) -> None:
        mocked_extract.return_value = "Insured: Unknown Person\nGrand Total: 5400"

        response = self.client.post(
            "/api/process",
            data={
                "workbook": (
                    io.BytesIO(self._workbook_bytes()),
                    "sales.xlsx",
                ),
                "insurance_files": (
                    io.BytesIO(b"insurance"),
                    "insurance-one.pdf",
                ),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        assert payload is not None

        self.assertEqual(payload["summary"]["billsUpdated"], 0)
        self.assertEqual(payload["summary"]["billsReview"], 1)
        self.assertEqual(len(payload["reviewRows"]), 1)
        self.assertEqual(payload["reviewRows"][0]["billFile"], "insurance-one.pdf")
        self.assertEqual(payload["reviewRows"][0]["reason"], "NO_MATCH")

    @patch("insurance_rto_updater.orchestration.pipeline.extract_text_from_file")
    def test_process_endpoint_preserves_existing_rto_values_when_clear_disabled(
        self,
        mocked_extract,
    ) -> None:
        mocked_extract.return_value = "Received From: Suresh Sharma\nGrand Total: 3200"

        response = self.client.post(
            "/api/process",
            data={
                "workbook": (
                    io.BytesIO(self._workbook_bytes(existing_rto=1800)),
                    "sales.xlsx",
                ),
                "rto_files": (
                    io.BytesIO(b"rto"),
                    "rto-one.pdf",
                ),
                "clear_existing": "0",
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        assert payload is not None

        self.assertEqual(payload["summary"]["billsUpdated"], 0)
        self.assertEqual(payload["summary"]["rowsUpdated"], 0)
        self.assertEqual(payload["summary"]["billsReview"], 1)
        self.assertEqual(payload["reviewRows"][0]["reason"], "EXISTING_TARGET_VALUE")
        self.assertEqual(payload["reviewRows"][0]["extractedAmount"], "3700")
        self.assertEqual(payload["rows"][1][2], 1800)

        download_response = self.client.get(payload["downloadUrl"])
        self.assertEqual(download_response.status_code, 200)

        workbook = load_workbook(io.BytesIO(download_response.data))
        worksheet = workbook[payload["sheetTitle"]]
        self.assertEqual(worksheet.cell(row=3, column=3).value, 1800)

    @patch("insurance_rto_updater.orchestration.pipeline.extract_text_from_file")
    def test_process_endpoint_uses_filename_first_name_fallback(
        self,
        mocked_extract,
    ) -> None:
        mocked_extract.return_value = "Grand Total: 5937"

        response = self.client.post(
            "/api/process",
            data={
                "workbook": (
                    io.BytesIO(
                        self._workbook_bytes(
                            customer_rows=[
                                ["VIRENDRA VALIYA S/O GIRDHARI", "", ""],
                                ["OTHER PERSON", "", ""],
                            ]
                        )
                    ),
                    "sales.xlsx",
                ),
                "rto_files": (
                    io.BytesIO(b"rto"),
                    "virendra_valiya_girdhari_2_feb.pdf",
                ),
                "clear_existing": "1",
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        assert payload is not None

        self.assertEqual(payload["summary"]["billsUpdated"], 1)
        self.assertEqual(payload["summary"]["parseFailures"], 0)
        self.assertEqual(payload["reviewRows"], [])
        self.assertEqual(payload["rows"][0][2], 6437.0)

    def test_process_endpoint_requires_receipts(self) -> None:
        response = self.client.post(
            "/api/process",
            data={
                "workbook": (
                    io.BytesIO(self._workbook_bytes()),
                    "sales.xlsx",
                ),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.get_json()["error"],
            "Upload at least one insurance or RTO bill.",
        )

    def test_generate_sales_register_endpoint_success(self) -> None:
        raw_data = [
            {
                "Invoice Date": "2026-02-14",
                "Invoice Number": "INV-001",
                "Contact Name": "John Doe",
                "Contact Mobile Number": "9999988888",
                "Contact Address": "123 Main St",
                "Model Name": "Model S",
                "Color": "Red",
                "VIN": "1234567890VIN",
                "Total Invoice Amount": 1500000.0,
                "Pre Vat Discount": 10000.0,
                "Financer Name": "HDFC",
                "Invoice Status": "Invoiced",
            }
        ]
        df = pd.DataFrame(raw_data)
        excel_io = io.BytesIO()
        df.to_excel(excel_io, index=False)
        excel_io.seek(0)

        response = self.client.post(
            "/api/generate-sales-register",
            data={
                "raw_file": (excel_io, "raw_invoices.xlsx"),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertIsNotNone(payload)
        self.assertEqual(payload["rowsWritten"], 1)
        self.assertEqual(payload["monthYear"], "Feb2026")
        self.assertIn("downloadUrl", payload)
        self.assertIn("manualColumns", payload)

    def test_generate_sales_register_endpoint_requires_file(self) -> None:
        response = self.client.post(
            "/api/generate-sales-register",
            data={},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("error", response.get_json())


if __name__ == "__main__":
    unittest.main()
