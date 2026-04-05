from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from insurance_rto_updater.integrations.local_workbook import (
    FIXED_CUSTOMER_HEADER,
    FIXED_INSURANCE_HEADER,
    FIXED_RTO_HEADER,
    LocalWorkbookAdapter,
)
from insurance_rto_updater.models import CellValueUpdate, SheetWritePlan


class LocalWorkbookAdapterTests(unittest.TestCase):
    def _make_workbook(self, path: Path) -> None:
        workbook = Workbook()
        ignored = workbook.active
        ignored.title = "Notes"
        ignored.append(["Other Header"])

        sales = workbook.create_sheet("Vehicle Sales Register")
        sales.append(
            [
                "Invoice No.",
                FIXED_CUSTOMER_HEADER,
                FIXED_INSURANCE_HEADER,
                FIXED_RTO_HEADER,
            ]
        )
        sales.append(["INV-1", "Ramesh Kumar", "", ""])
        sales.append(["INV-2", "Suresh Sharma", 1200, 1800])
        workbook.save(path)

    def test_load_sheet_data_uses_first_matching_worksheet(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "sales.xlsx"
            self._make_workbook(workbook_path)

            adapter = LocalWorkbookAdapter(workbook_path)
            sheet_data = adapter.load_sheet_data()

            self.assertEqual(sheet_data.sheet_title, "Vehicle Sales Register")
            self.assertEqual(sheet_data.header_row[1], FIXED_CUSTOMER_HEADER)
            self.assertEqual(sheet_data.data_rows[0][1], "Ramesh Kumar")

    def test_apply_write_plan_clears_and_updates_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "sales.xlsx"
            self._make_workbook(workbook_path)

            adapter = LocalWorkbookAdapter(workbook_path)
            sheet_data = adapter.load_sheet_data()

            adapter.apply_write_plan(
                sheet_title=sheet_data.sheet_title,
                write_plan=SheetWritePlan(
                    clear_from_row=2,
                    clear_to_row=3,
                    clear_columns=[3, 4],
                    value_updates=[
                        CellValueUpdate(row_index=2, col_index=3, value=5400.0),
                        CellValueUpdate(row_index=3, col_index=4, value=3200.0),
                    ],
                ),
            )
            updated_path = adapter.save_copy(Path(temp_dir) / "sales_updated.xlsx")
            updated = load_workbook(updated_path)
            worksheet = updated[sheet_data.sheet_title]

            self.assertEqual(worksheet.cell(row=2, column=3).value, 5400.0)
            self.assertEqual(worksheet.cell(row=3, column=4).value, 3200.0)
            self.assertIsNone(worksheet.cell(row=2, column=4).value)
            self.assertIsNone(worksheet.cell(row=3, column=3).value)

    def test_load_sheet_data_raises_when_headers_are_missing(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "sales.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            worksheet.append(["Customer Name", "Insurance", "RTO"])
            workbook.save(workbook_path)

            adapter = LocalWorkbookAdapter(workbook_path)
            with self.assertRaises(ValueError) as context:
                adapter.load_sheet_data()

            self.assertIn("Workbook must contain a worksheet", str(context.exception))


if __name__ == "__main__":
    unittest.main()
