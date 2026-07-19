from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from insurance_rto_updater.output.sales_register import generate_sales_register


class TestSalesRegister(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.temp_path = Path(self.temp_dir.name)

    def test_generate_sales_register_success(self) -> None:
        # Create a mock raw input Excel sheet with pandas
        raw_data = [
            {
                "Invoice Date": "2026-02-14",
                "Invoice Number": "INV-001",
                "Contact Name": "John Doe",
                "Contact Mobile Number": "9999988888",
                "Contact Address": "123 Main St",
                "Model Name": "Model S",
                "Color": "Red",
                "VIN": "1234567890VIN",
                "Total Invoice Amount": 1500000.0,
                "Pre Vat Discount": 10000.0,
                "Financer Name": "HDFC",
                "Invoice Status": "Invoiced",
            },
            {
                "Invoice Date": "2026-02-15",
                "Invoice Number": "INV-002",
                "Contact Name": "Jane Smith",
                "Contact Mobile Number": "8888877777",
                "Contact Address": "456 Side St",
                "Model Name": "Model X",
                "Color": "Blue",
                "VIN": "0987654321VIN",
                "Total Invoice Amount": 1800000.0,
                "Pre Vat Discount": 0.0,
                "Financer Name": None,
                "Invoice Status": "Draft",  # This row should be filtered out
            },
        ]
        df = pd.DataFrame(raw_data)
        input_excel = self.temp_path / "raw_invoices.xlsx"
        df.to_excel(input_excel, index=False)

        output_excel = self.temp_path / "output_register.xlsx"
        result = generate_sales_register(input_excel, output_excel)

        # Check result metadata
        self.assertEqual(result.rows_written, 1)
        self.assertEqual(result.month_year, "Feb2026")
        self.assertTrue(output_excel.exists())

        # Load generated workbook to verify contents
        wb = load_workbook(output_excel, data_only=False)
        self.assertIn("Vehicle Sales Register", wb.sheetnames)
        ws = wb["Vehicle Sales Register"]

        # Verify header count
        self.assertEqual(ws.max_column, 28)

        # Verify data rows (row 2 should correspond to INV-001)
        self.assertEqual(ws["B2"].value, "INV-001")
        self.assertEqual(ws["C2"].value, "John Doe")
        self.assertEqual(ws["F2"].value, "Model S")
        self.assertEqual(ws["H2"].value, "1234567890VIN")
        self.assertEqual(ws["J2"].value, 1500000.0)

        # Check ex-showroom formula K2: "=J2-1099.72"
        self.assertEqual(ws["K2"].value, "=J2-1099.72")

        # Check discount calculation: 10000 * 1.18 = 11800
        self.assertEqual(ws["O2"].value, 11800.0)

        # Check HP HP-related formulas
        self.assertEqual(ws["U2"].value, "=IF(ISTEXT(Z2),1000,0)")

        # Verify filtered out row is not present (max row should be 2)
        self.assertEqual(ws.max_row, 2)
